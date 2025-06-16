# File: processors/transcript_processor_full.py
"""
Full transcript processing module integrating all features from auto-meeting-minutes:
- VTT/SRT to TXT conversion
- TXT to XLSX with speaker colors and time gradients
- Timestamp refinement with NLP
- Excel to HTML with AI summaries
- Enhanced speaker summaries with multiple topics
"""

import os
import re
import json
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, Any, Optional, List, Tuple
from datetime import datetime
import colorsys
import random
import hashlib

from processors.base_processor import BaseProcessor
from core.config import Config
from core.utils import sanitize_filename
from core.prompts import PromptManager

# For Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# For NLP and refinement
try:
    import nltk
    nltk.download('punkt_tab')
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity
    NLTK_AVAILABLE = True
    
    # Download required NLTK data
    try:
        nltk.data.find('tokenizers/punkt')
    except LookupError:
        print("Downloading NLTK punkt tokenizer...")
        nltk.download('punkt')
    
    try:
        nltk.data.find('corpora/stopwords')
    except LookupError:
        print("Downloading NLTK stopwords...")
        nltk.download('stopwords')
        
except ImportError:
    NLTK_AVAILABLE = False
    print("Warning: NLTK/sklearn not installed. Timestamp refinement will be limited.")

# For AI summaries
try:
    import openai
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False
    print("Warning: OpenAI not installed. Summaries will not be generated.")


class TranscriptProcessor(BaseProcessor):
    """Full-featured transcript processor with all original functionality."""
    
    def __init__(self, output_dir: Path):
        super().__init__(output_dir)
        self.metadata = {}
        self.api_key = Config.OPENAI_API_KEY
        self.gpt_model = Config.GPT_MODEL
        self.batch_size_minutes = Config.BATCH_SIZE_MINUTES
        # Initialize prompt manager
        custom_prompts_path = Path(os.getenv("GPT_PROMPTS_FILE", "prompts.json"))
        self.prompt_manager = PromptManager(custom_prompts_path)
    
    def process(self, transcript_file: Path, video_id: Optional[str] = None, **kwargs) -> Dict[str, Any]:
        """Process VTT/SRT transcript through the complete pipeline."""
        try:
            # Store metadata for use in other methods
            self.metadata = kwargs.get('metadata', {})
            print(f"\nStarting full transcript processing pipeline...")
            
            # Step 1: Convert VTT/SRT to TXT
            print("Step 1: Converting subtitle to text...")
            txt_file = self._vtt_to_txt(transcript_file)
            
            # Step 2: Convert TXT to XLSX with full formatting
            print("Step 2: Converting text to Excel with formatting...")
            xlsx_file = self._txt_to_xlsx_full(txt_file)
            
            # Step 3: Refine timestamps (optional)
            if not kwargs.get('skip_refinement', False) and NLTK_AVAILABLE:
                print("Step 3: Refining timestamps...")
                refined_xlsx = self._refine_timestamps_full(xlsx_file)
                if refined_xlsx:
                    xlsx_file = refined_xlsx
            else:
                print("Step 3: Skipping timestamp refinement...")
            
            # Step 4: Generate HTML and summaries with all features
            print("Step 4: Generating HTML and summaries...")
            html_files = self._xlsx_to_html_full(xlsx_file, video_id, **kwargs)
            
            # Step 5: Convert markdown bold to HTML (optional)
            if not kwargs.get('skip_bold_conversion', False):
                print("Step 5: Converting markdown bold formatting...")
                self._convert_bold_formatting(html_files)
            
            return {
                'success': True,
                'files': {
                    'txt': str(txt_file),
                    'xlsx': str(xlsx_file),
                    **html_files
                },
                'metadata': {
                    'processed': True,
                    'has_video_id': video_id is not None,
                    'summaries_generated': OPENAI_AVAILABLE and self.api_key is not None,
                    'refinement_applied': not kwargs.get('skip_refinement', False) and NLTK_AVAILABLE
                }
            }
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return {
                'success': False,
                'error': f'Transcript processing failed: {str(e)}'
            }
    
    def _vtt_to_txt(self, vtt_file: Path) -> Path:
        """Convert VTT/SRT to plain text transcript with proper speaker formatting."""
        # Get meeting name from metadata or use file stem
        meeting_name = self.metadata.get('meeting_name', vtt_file.stem) if hasattr(self, 'metadata') else vtt_file.stem
        output_file = self.output_dir / f"{meeting_name}.txt"
        
        # Regular expression to match timestamp, speaker, and text
        pattern = r'(\d{2}:\d{2}:\d{2})[,.]?\d{3}\s*-->'
        
        # Read the file
        with open(vtt_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        output_lines = []
        
        # Skip WEBVTT header if present
        start_index = 0
        if lines and "WEBVTT" in lines[0]:
            start_index = 1
        
        # Process the content
        i = start_index
        while i < len(lines):
            line = lines[i].strip()
            
            # Skip empty lines
            if not line:
                i += 1
                continue
            
            # Check if line is a cue number (just a number)
            if re.match(r'^\d+$', line):
                # Next line should be the timestamp
                i += 1
                if i < len(lines):
                    timestamp_line = lines[i].strip()
                    # Look for timestamps
                    timestamp_match = re.search(pattern, timestamp_line)
                    
                    if timestamp_match:
                        # Get the full time part (HH:MM:SS)
                        full_time = timestamp_match.group(1)
                        
                        # Move to the next line which contains the text content
                        i += 1
                        if i < len(lines):
                            text_line = lines[i].strip()
                            if text_line:
                                # # Format as "HH:MM:SS Speaker: Text"
                                # # If no speaker format detected, add a default
                                # if ': ' not in text_line:
                                #     text_line = f"Speaker: {text_line}"
                                # output_lines.append(f"{full_time} {text_line}")
                                # Handle both [SPEAKER]: and SPEAKER: formats
                                speaker_match = re.match(r'^\[?([^:\]]+)\]?:\s*(.+)', text_line)
                                
                                if speaker_match:
                                    speaker = speaker_match.group(1).strip()
                                    text = speaker_match.group(2).strip()
                                    output_lines.append(f"{full_time} {speaker}: {text}")
                                else:
                                    # No speaker format detected, add default
                                    output_lines.append(f"{full_time} Speaker: {text_line}")
            
            # Move to next line
            i += 1
        
        # Write the formatted transcript
        with open(output_file, 'w', encoding='utf-8') as f:
            for line in output_lines:
                f.write(f"{line}\n")
        
        print(f"  Created text transcript: {output_file}")
        return output_file
    
    def _txt_to_xlsx_full(self, txt_file: Path) -> Path:
        """Convert text to Excel with full formatting including rainbow gradients."""
        # Get meeting name from metadata or use file stem
        meeting_name = self.metadata.get('meeting_name', txt_file.stem) if hasattr(self, 'metadata') else txt_file.stem
        output_file = self.output_dir / f"{meeting_name}.xlsx"
        
        # Pattern to match timestamp, speaker, and text
        # pattern = r'(\d{2}:\d{2}:\d{2}) ([^:]+): (.+)'
        pattern = r'(\d{2}:\d{2}:\d{2}) \[?([^:\]]+)\]?:\s*(.+)'
        
        # Read transcript
        with open(txt_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Extract data
        matches = re.findall(pattern, content)
        data = []
        
        # Track first occurrence of each speaker
        first_occurrences = {}
        
        # Collect all unique speakers first
        all_speakers = set()
        for time_str, speaker, text in matches:
            all_speakers.add(speaker)
        
        # Generate unique colors for all speakers
        speaker_colors = self._generate_unique_colors(all_speakers)
        
        for time_str, speaker, text in matches:
            seconds = self._time_to_seconds(time_str)
            
            # Check if this is the first occurrence of the speaker
            first_time = None
            first_seconds = None
            first_speaker = None
            
            if speaker not in first_occurrences:
                first_occurrences[speaker] = (time_str, seconds)
                first_time = time_str
                first_seconds = seconds
                first_speaker = speaker
            
            data.append({
                'Seconds': seconds,
                'Time': time_str,
                'First': first_speaker,
                'First_Time': first_time,
                'First_Seconds': first_seconds,
                'Name': speaker,
                'Text': text
            })
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Add topic tracking columns
        df['Topic_Number'] = None
        df['Topic_Start_Time'] = None
        df['Topic_Start_Seconds'] = None
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Meeting Transcript"
        
        # Add headers
        headers = [
            'Seconds', 'Time', 'First', 'First_Time', 'First_Seconds', 
            'Name', 'Text', 'Topic_Number', 'Topic_Start_Time', 'Topic_Start_Seconds'
        ]
        
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num).value = header
            ws.cell(row=1, column=col_num).font = Font(bold=True)
        
        # Calculate gradient positions for rainbow effect
        if data:
            min_seconds = min(row['Seconds'] for row in data)
            max_seconds = max(row['Seconds'] for row in data)
            time_range = max_seconds - min_seconds if max_seconds > min_seconds else 1
        else:
            time_range = 1
        
        # Add data and apply formatting
        for row_num, row_data in enumerate(data, 2):
            # Calculate time gradient position (0-1)
            time_position = (row_data['Seconds'] - min_seconds) / time_range if time_range > 0 else 0
            
            # Get rainbow color for time
            time_color = self._get_rainbow_color(time_position)
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                
                # Get the value from row_data or DataFrame
                if header in row_data:
                    cell.value = row_data[header]
                else:
                    # Get from DataFrame for added columns
                    cell.value = df.iloc[row_num-2][header] if header in df.columns else None
                
                # Apply rainbow gradient to time-based columns
                if header in ('Seconds', 'First_Seconds', 'Topic_Start_Seconds'):
                    if cell.value is not None:
                        cell.fill = PatternFill(start_color=time_color, end_color=time_color, fill_type="solid")
                
                # Apply color to speaker names (except specific exclusions)
                elif header in ('Name', 'First') and cell.value and cell.value != "Manolis Kellis":
                    speaker = cell.value
                    if speaker in speaker_colors:
                        cell.fill = PatternFill(start_color=speaker_colors[speaker], 
                                               end_color=speaker_colors[speaker],
                                               fill_type="solid")
        
        # Auto-adjust column width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 100)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(output_file)
        
        print(f"  Created Excel file: {output_file}")
        print(f"  Total entries: {len(data)}")
        print(f"  Unique speakers: {len(all_speakers)}")
        
        return output_file
    
    def _refine_timestamps_full(self, xlsx_file: Path) -> Optional[Path]:
        """Refine timestamps using NLP for better topic matching."""
        if not NLTK_AVAILABLE:
            return None
        
        try:
            from nltk.tokenize import word_tokenize
            from nltk.corpus import stopwords
            
            # Get meeting name from metadata or use file stem
            meeting_name = self.metadata.get('meeting_name', xlsx_file.stem) if hasattr(self, 'metadata') else xlsx_file.stem
            output_file = self.output_dir / f"{meeting_name}_refined.xlsx"

            # Read Excel file
            df = pd.read_excel(xlsx_file)

            # Ensure required columns exist with correct types
            if 'Topic_Start_Time' not in df.columns:
                df['Topic_Start_Time'] = pd.Series(dtype=str)
            df['Topic_Start_Time'] = df['Topic_Start_Time'].astype(str)

            if 'Topic_Number' not in df.columns:
                df['Topic_Number'] = pd.Series(dtype='Int64')

            if 'Topic_Start_Seconds' not in df.columns:
                df['Topic_Start_Seconds'] = pd.Series(dtype='Int64')
            
            # Extract transcript entries
            transcript_entries = []
            for i, row in df.iterrows():
                if pd.notna(row['Name']) and pd.notna(row['Seconds']) and pd.notna(row['Text']):
                    entry = {
                        'name': row['Name'],
                        'seconds': int(row['Seconds']),
                        'time_str': row['Time'],
                        'text': row['Text'],
                        'row_index': i
                    }
                    transcript_entries.append(entry)
            
            # Detect topic changes for each speaker
            speaker_topics = self._detect_speaker_topics(transcript_entries)
            
            # Update DataFrame with topic information
            for speaker, topics in speaker_topics.items():
                for topic_num, topic in enumerate(topics, 1):
                    for row_idx in topic['indices']:
                        df.loc[row_idx, 'Topic_Number'] = topic_num
                        df.loc[row_idx, 'Topic_Start_Time'] = str(topic['start_time'])
                        df.loc[row_idx, 'Topic_Start_Seconds'] = topic['start_seconds']
            
            # Save refined Excel file
            df.to_excel(output_file, index=False)
            
            print(f"  Refined timestamps saved to: {output_file}")
            return output_file
            
        except Exception as e:
            print(f"  Warning: Timestamp refinement failed: {e}")
            return None
    
    def _xlsx_to_html_full(self, xlsx_file: Path, video_id: Optional[str] = None, **kwargs) -> Dict[str, str]:
        """Generate full HTML and markdown outputs with all features."""
        # Get metadata from kwargs or instance
        metadata = kwargs.get('metadata', getattr(self, 'metadata', {}))
        files = {}
        
        # Read Excel file
        df = pd.read_excel(xlsx_file)
        
        # Extract transcript data
        transcript_data = []
        for i, row in df.iterrows():
            if pd.notna(row['Name']) and pd.notna(row['Seconds']) and pd.notna(row['Text']):
                transcript_data.append({
                    'name': row['Name'],
                    'seconds': int(row['Seconds']),
                    'time_str': row['Time'],
                    'text': row['Text'],
                    'row_index': i
                })
        
        # Check if we should use enhanced summaries
        use_enhanced = kwargs.get('use_enhanced_summaries', True)
        
        if use_enhanced and OPENAI_AVAILABLE and self.api_key:
            print("  Using enhanced speaker summaries with multiple topics...")
            
            # Generate enhanced summaries data
            summaries_data = self._generate_speaker_summaries_data(transcript_data)
            
            # Generate enhanced HTML
            speaker_html = self._generate_enhanced_speaker_html(transcript_data, video_id, summaries_data, metadata)
            meeting_name = metadata.get('meeting_name', xlsx_file.stem) if metadata else xlsx_file.stem
            speaker_html_file = self.output_dir / f"{meeting_name}_speaker_summaries.html"
            with open(speaker_html_file, 'w', encoding='utf-8') as f:
                f.write(speaker_html)
            files['speaker_html'] = str(speaker_html_file)
            
            # Generate enhanced markdown
            speaker_md = self._generate_enhanced_speaker_markdown(transcript_data, video_id, summaries_data, metadata)
            speaker_md_file = self.output_dir / f"{meeting_name}_speaker_summaries.md"
            with open(speaker_md_file, 'w', encoding='utf-8') as f:
                f.write(speaker_md)
            files['speaker_md'] = str(speaker_md_file)
        else:
            # Generate basic speaker links
            speaker_html = self._generate_basic_speaker_html(df, video_id)
            meeting_name = metadata.get('meeting_name', xlsx_file.stem) if metadata else xlsx_file.stem
            speaker_html_file = self.output_dir / f"{meeting_name}_speaker_summaries.html"
            with open(speaker_html_file, 'w', encoding='utf-8') as f:
                f.write(speaker_html)
            files['speaker_html'] = str(speaker_html_file)
        
        # Generate meeting summaries if API available
        if OPENAI_AVAILABLE and self.api_key:
            print("  Generating meeting summaries...")
            meeting_html, meeting_md = self._generate_meeting_summaries_full(transcript_data, video_id)
            
            meeting_html_file = self.output_dir / f"{meeting_name}_meeting_summaries.html"
            with open(meeting_html_file, 'w', encoding='utf-8') as f:
                f.write(meeting_html)
            files['meeting_html'] = str(meeting_html_file)
            
            meeting_md_file = self.output_dir / f"{meeting_name}_meeting_summaries.md"
            with open(meeting_md_file, 'w', encoding='utf-8') as f:
                f.write(meeting_md)
            files['meeting_md'] = str(meeting_md_file)
        
        print(f"  Generated output files in: {self.output_dir}")
        return files
    
    # Helper methods
    
    def _time_to_seconds(self, time_str: str) -> int:
        """Convert HH:MM:SS to seconds."""
        parts = time_str.split(':')
        return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
    
    def _seconds_to_time_str(self, seconds: int) -> str:
        """Convert seconds to HH:MM:SS format."""
        hours = seconds // 3600
        minutes = (seconds % 3600) // 60
        secs = seconds % 60
        return f"{hours}:{minutes:02d}:{secs:02d}"
    
    def _generate_unique_colors(self, speakers: set) -> Dict[str, str]:
        """Generate unique colors for each speaker using HSV color space."""
        # Filter out specific speakers if needed
        filtered_speakers = [s for s in speakers if s != "Manolis Kellis"]
        
        colors = {}
        # Use golden ratio for optimal color distribution
        golden_ratio_conjugate = 0.618033988749895
        
        # Start at a random hue
        h = random.random()
        
        for speaker in filtered_speakers:
            # Higher saturation and value for lighter, more readable colors
            s = 0.4  # Lower saturation for lighter colors
            v = 0.95  # High value for brightness
            
            # Convert HSV to RGB
            r, g, b = [int(x * 255) for x in colorsys.hsv_to_rgb(h, s, v)]
            colors[speaker] = f"{r:02x}{g:02x}{b:02x}"
            
            # Increment hue by golden ratio for optimal spacing
            h += golden_ratio_conjugate
            h %= 1  # Keep within [0, 1]
        
        return colors
    
    def _get_rainbow_color(self, position: float) -> str:
        """Generate a color from rainbow gradient based on position (0-1)."""
        # Using HSV color space for rainbow effect
        h = position * 0.8  # Full spectrum
        s = 0.7  # Saturation
        v = 0.9  # Value
        
        r, g, b = [int(x * 255) for x in colorsys.hsv_to_rgb(h, s, v)]
        return f"{r:02x}{g:02x}{b:02x}"
    
    def _detect_speaker_topics(self, transcript_data: List[Dict]) -> Dict[str, List[Dict]]:
        """Detect topic changes for each speaker."""
        speaker_topics = {}
        current_topics = {}
        
        # Sort by timestamp
        sorted_data = sorted(transcript_data, key=lambda x: x['seconds'])
        
        for i, entry in enumerate(sorted_data):
            speaker = entry['name']
            seconds = entry['seconds']
            
            # Initialize speaker if first time seen
            if speaker not in speaker_topics:
                speaker_topics[speaker] = []
                current_topics[speaker] = {
                    'start_idx': i,
                    'start_time': entry['time_str'],
                    'start_seconds': seconds,
                    'text': [entry['text']],
                    'indices': [entry['row_index']]
                }
            else:
                # Check if this might be a new topic (5 minute gap)
                last_idx = current_topics[speaker]['indices'][-1]
                time_gap = seconds - sorted_data[last_idx]['seconds'] if last_idx < len(sorted_data) else 0
                
                if time_gap > 300:  # 5 minutes
                    # Finalize current topic
                    speaker_topics[speaker].append(current_topics[speaker])
                    
                    # Start new topic
                    current_topics[speaker] = {
                        'start_idx': i,
                        'start_time': entry['time_str'],
                        'start_seconds': seconds,
                        'text': [entry['text']],
                        'indices': [entry['row_index']]
                    }
                else:
                    # Continue current topic
                    current_topics[speaker]['text'].append(entry['text'])
                    current_topics[speaker]['indices'].append(entry['row_index'])
        
        # Add final topics
        for speaker, topic in current_topics.items():
            if topic['text']:
                speaker_topics[speaker].append(topic)
        
        return speaker_topics
    
    def _generate_speaker_summaries_data(self, transcript_data: List[Dict]) -> Dict[str, List[Dict]]:
        """Generate speaker summaries data with AI using customizable prompts."""
        if not OPENAI_AVAILABLE or not self.api_key:
            return {}
        
        topic_data = self._detect_speaker_topics(transcript_data)
        
        from openai import OpenAI
        client = OpenAI(api_key=self.api_key)
        
        for speaker, topics in topic_data.items():
            for i, topic in enumerate(topics, 1):
                topic_text = ' '.join(topic['text'])
                
                try:
                    # Get customized prompt
                    prompt = self.prompt_manager.get_prompt(
                        "SPEAKER_SUMMARY",
                        speaker=speaker,
                        topic_num=i,
                        transcript=topic_text
                    )
                    
                    # Get system prompt
                    system_prompt = self.prompt_manager.get_prompt("SYSTEM_SPEAKER")
                    
                    response = client.chat.completions.create(
                        model=self.gpt_model,
                        messages=[
                            {"role": "system", "content": system_prompt},
                            {"role": "user", "content": prompt}
                        ],
                        response_format={"type": "json_object"},
                        max_tokens=800,
                    )
                    
                    summary_json = json.loads(response.choices[0].message.content)
                    
                    topic['summary'] = {
                        'title': summary_json.get('title', f'Topic {i}'),
                        'content': summary_json.get('content', topic_text[:100] + '...')
                    }
                    
                except Exception as e:
                    print(f"Warning: Failed to generate summary for {speaker} topic {i}: {e}")
                    topic['summary'] = {
                        'title': f'Topic {i}',
                        'content': topic_text[:100] + '...'
                    }
        
        return topic_data
    
    def _generate_enhanced_speaker_html(self, transcript_data: List[Dict], video_id: Optional[str], 
                                            summaries_data: Dict[str, List[Dict]], metadata: Dict = None) -> str:
        """Generate enhanced HTML with speaker summaries and proper title."""
        html_content = '<!DOCTYPE html>\n<html>\n<head>\n<title>Speaker Summaries</title>\n'
        html_content += '<style>\n'
        html_content += 'body { font-family: Arial, sans-serif; margin: 20px; font-size: 11pt; }\n'
        html_content += 'h1 { font-family: Cambria, serif; font-size: 11pt; color: #c0504d; text-decoration: underline; margin-bottom: 0px; display: inline-block; }\n'
        html_content += 'h1 a { color: #c0504d; text-decoration: underline; }\n'
        html_content += '.speaker { font-weight: bold; color: #7030a0; text-decoration: underline; margin-bottom: 3px; }\n'
        html_content += '.topic { margin-left: 0px; margin-bottom: 3px; }\n'
        html_content += '.topic-title { font-weight: bold; color: #1f497d; text-decoration: underline; }\n'
        html_content += 'ol { list-style-position: outside; padding-left: 12px; margin-top: 0px; }\n'
        html_content += 'ol li { margin-bottom: 0px; }\n'
        html_content += 'a { color: inherit; text-decoration: none; }\n'
        html_content += '.timestamp { color: #1155cc; }\n'
        html_content += "b { font-weight: bold; }\n"
        html_content += '</style>\n</head>\n<body>\n'
        
        # Add title with meeting name
        meeting_name = metadata.get('meeting_name', 'Meeting') if metadata else 'Meeting'
        formatted_name = self._format_meeting_name(meeting_name)

        # Add title
        if video_id:
            video_link = f'https://mit.hosted.panopto.com/Panopto/Pages/Viewer.aspx?id={video_id}'
            html_content += f'<h1><a href="{video_link}">{formatted_name} <span style="color: #1155cc;">(link)</span></a></h1>\n'
        else:
            html_content += f'<h1>{formatted_name}</h1>\n'
        
        # Create an ordered list for speakers
        html_content += '<ol>\n'
        
        # Process each speaker
        for speaker_idx, (speaker, topics) in enumerate(summaries_data.items(), 1):
            # Speaker name as a list item
            html_content += f'<li><div class="speaker">{speaker}</div>\n'
            
            # Process each topic
            for i, topic in enumerate(topics, 1):
                # Get the summary
                topic_summary = topic.get('summary', {'title': f'Topic {i}', 'content': 'Summary not available'})
                
                # Format timestamp link
                timestamp_seconds = topic['start_seconds']
                timestamp_str = topic['start_time']
                
                if video_id:
                    video_link = f'https://mit.hosted.panopto.com/Panopto/Pages/Viewer.aspx?id={video_id}&start={timestamp_seconds}'
                    html_content += f'<div class="topic">(<span class="topic-title">{i}) {topic_summary["title"]}</span> '
                    html_content += f'<a href="{video_link}"><span class="timestamp">({timestamp_str})</span></a>: '
                    html_content += f'{topic_summary["content"]}</div>\n'
                else:
                    html_content += f'<div class="topic">(<span class="topic-title">{i}) {topic_summary["title"]}</span> '
                    html_content += f'<span class="timestamp">({timestamp_str})</span>: '
                    html_content += f'{topic_summary["content"]}</div>\n'
            
            html_content += '</li>\n'
        
        html_content += '</ol>\n</body>\n</html>'
        
        return html_content
    
    def _generate_enhanced_speaker_markdown(self, transcript_data: List[Dict], video_id: Optional[str],
                                               summaries_data: Dict[str, List[Dict]], metadata: Dict = None) -> str:
        """Generate enhanced markdown with speaker summaries and proper title."""
        md_lines = []
        
        # Add title with meeting name
        meeting_name = metadata.get('meeting_name', 'Meeting') if metadata else 'Meeting'
        formatted_name = self._format_meeting_name(meeting_name)
        
        if video_id:
            video_link = f"https://mit.hosted.panopto.com/Panopto/Pages/Viewer.aspx?id={video_id}"
            md_lines.append(f"# [{formatted_name}]({video_link})\n")
        else:
            md_lines.append(f"# {formatted_name}\n")
        
        # Process each speaker
        for speaker, topics in summaries_data.items():
            # Speaker name as header
            md_lines.append(f"**{speaker}**")
            
            # Process each topic
            for i, topic in enumerate(topics, 1):
                # Get the summary
                topic_summary = topic.get('summary', {'title': f'Topic {i}', 'content': 'Summary not available'})
                
                # Format timestamp link
                timestamp_seconds = topic['start_seconds']
                timestamp_str = topic['start_time']
                
                if video_id:
                    video_link = f'https://mit.hosted.panopto.com/Panopto/Pages/Viewer.aspx?id={video_id}&start={timestamp_seconds}'
                    md_lines.append(f"**({i}) {topic_summary['title']} **[({timestamp_str})]({video_link}): {topic_summary['content']}")
                else:
                    md_lines.append(f"**({i}) {topic_summary['title']} **({timestamp_str}): {topic_summary['content']}")
            
            # Add blank line between speakers
            if speaker != list(summaries_data.keys())[-1]:
                md_lines.append("")
        
        return '\n'.join(md_lines)
    
    def _generate_basic_speaker_html(self, df: pd.DataFrame, video_id: Optional[str]) -> str:
        """Generate basic speaker links HTML (fallback when no AI available)."""
        html = """<!DOCTYPE html>
<html>
<head>
<title>Speaker Links</title>
<style>
body { font-family: Arial, sans-serif; margin: 20px; }
.speaker { margin: 10px 0; }
.speaker-name { font-weight: bold; color: #7030a0; }
.timestamp { color: #1155cc; }
a { text-decoration: none; }
a:hover { text-decoration: underline; }
</style>
</head>
<body>
<h1>Speaker Links</h1>
"""
        
        # Get unique speakers
        if 'First' in df.columns:
            speakers = df[df['First'].notna()][['First', 'First_Time', 'First_Seconds']].values
        else:
            speakers = []
            seen = set()
            for _, row in df.iterrows():
                if row['Name'] not in seen:
                    seen.add(row['Name'])
                    speakers.append([row['Name'], row['Time'], row['Seconds']])
        
        # Generate links
        for speaker, time_str, seconds in speakers:
            if video_id:
                link = f'https://mit.hosted.panopto.com/Panopto/Pages/Viewer.aspx?id={video_id}&start={int(seconds)}'
                html += f'<div class="speaker"><span class="speaker-name">{speaker}</span> '
                html += f'<a href="{link}"><span class="timestamp">({time_str})</span></a></div>\n'
            else:
                html += f'<div class="speaker"><span class="speaker-name">{speaker}</span> '
                html += f'<span class="timestamp">({time_str})</span></div>\n'
        
        html += "</body>\n</html>"
        
        return html
    
    def _generate_meeting_summaries_full(self, transcript_data: List[Dict], video_id: Optional[str]) -> Tuple[str, str]:
        """Generate meeting summaries with batching and AI."""
        if not OPENAI_AVAILABLE or not self.api_key:
            return self._generate_meeting_summaries_fallback()
        
        # Create time-based batches
        batches = self._create_time_batches(transcript_data)
        print(f"    Created {len(batches)} batches for summarization")
        
        # Generate summaries for each batch
        batch_summaries = []
        for i, batch in enumerate(batches, 1):
            print(f"    Processing batch {i}/{len(batches)}...")
            summary = self._summarize_batch(batch, i)
            batch_summaries.append(summary)
        
        # Generate HTML
        html = self._generate_meeting_html(batches, batch_summaries, video_id, transcript_data, self.metadata)
        
        # Generate Markdown
        md = self._generate_meeting_markdown(batches, batch_summaries, video_id, transcript_data, self.metadata)
        
        return html, md
    
    def _create_time_batches(self, transcript_data: List[Dict]) -> List[List[Dict]]:
        """Create time-based batches for summarization."""
        if not transcript_data:
            return []
        
        # Sort by timestamp
        sorted_data = sorted(transcript_data, key=lambda x: x['seconds'])
        
        # Get start and end time
        start_time = sorted_data[0]['seconds']
        end_time = sorted_data[-1]['seconds']
        
        # Convert batch size to seconds
        batch_size_seconds = self.batch_size_minutes * 60
        
        # Create batches
        batches = []
        batch_start = start_time
        
        while batch_start < end_time:
            batch_end = min(batch_start + batch_size_seconds, end_time)
            
            # Get entries for this batch
            batch_entries = [
                entry for entry in sorted_data
                if batch_start <= entry['seconds'] < batch_end
            ]
            
            if batch_entries:
                batches.append(batch_entries)
            
            batch_start = batch_end
        
        return batches
    
    def _summarize_batch(self, batch_entries: List[Dict], batch_number: int) -> str:
        """Summarize a batch using customizable prompts."""
        if not self.api_key:
            return "API key not provided. Summaries not generated."
        
        # Prepare batch text and timestamps
        batch_text = ""
        for entry in sorted(batch_entries, key=lambda x: x['seconds']):
            batch_text += f"{entry['name']}: {entry['text']}\n\n"
        
        # Get time bounds
        start_seconds = min(entry['seconds'] for entry in batch_entries)
        end_seconds = max(entry['seconds'] for entry in batch_entries)
        start_time = self._seconds_to_time_str(start_seconds)
        end_time = self._seconds_to_time_str(end_seconds)
        
        # Create timestamp reference
        timestamp_reference = self._create_timestamp_reference(batch_entries)
        
        try:
            from openai import OpenAI
            client = OpenAI(api_key=self.api_key)
            
            # Get customized prompt
            prompt = self.prompt_manager.get_prompt(
                "BATCH_SUMMARY",
                timestamp_reference=timestamp_reference,
                batch_number=batch_number,
                start_time=start_time,
                end_time=end_time,
                transcript=batch_text
            )
            
            # Get system prompt
            system_prompt = self.prompt_manager.get_prompt("SYSTEM_BATCH")
            
            response = client.chat.completions.create(
                model=self.gpt_model,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=10000,
            )
            
            return response.choices[0].message.content.strip()
            
        except Exception as e:
            return f"Error generating batch summary: {str(e)}"
    
    def _generate_meeting_html(self, batches: List[List[Dict]], batch_summaries: List[str], 
                          video_id: Optional[str], transcript_data: List[Dict], 
                          metadata: Dict = None) -> str:
        """Generate HTML for meeting summaries - matching original formatting."""
        # Start HTML with exact original styling
        html_content = "<!DOCTYPE html>\n<html>\n<head>\n<title>Meeting Summaries</title>\n"
        html_content += "<style>\n"
        html_content += "body { font-family: Arial, sans-serif; margin: 20px; font-size: 11pt; }\n"
        html_content += "ol { list-style-position: outside; padding-left: 12px; margin-top: 0px; }\n"
        html_content += "ol li { margin-bottom: 1px; }\n"  # Changed from 15px to 1px to match original
        html_content += ".topic-content { margin-bottom: 0px; font-family: Arial, sans-serif; font-size: 11pt; margin-top: 0px; }\n"
        # Title styling - Cambria, 11pt, #c0504d, underlined
        html_content += "h1 { font-family: Cambria, serif; font-size: 11pt; color: #c0504d; text-decoration: underline; display: inline-block; }\n"
        html_content += "h1 a { color: #c0504d; text-decoration: underline; }\n"
        # Topic styling - Arial, 11pt, #7030a0, underlined
        html_content += "h3.topic-heading { font-family: Arial, sans-serif; font-size: 11pt; color: #7030a0; text-decoration: underline; margin-top: 0px; margin-bottom: 1px; }\n"
        html_content += "a { color: inherit; }\n"
        html_content += ".topic-link { text-decoration: underline; color: #7030a0; }\n"
        html_content += ".topic-link span { text-decoration: underline; }\n"
        html_content += "b { font-weight: bold; }\n"
        html_content += "</style>\n</head>\n<body>\n"
        
        # Get meeting name from metadata
        meeting_name = metadata.get('meeting_name') if metadata else None
        if not meeting_name and hasattr(self, 'metadata'):
            meeting_name = self.metadata.get('meeting_name')
        if meeting_name:
            formatted_name = self._format_meeting_name(meeting_name)
        else:
            formatted_name = 'Meeting Summary'
        
        # Add title with exact original format
        if video_id:
            video_link = f"https://mit.hosted.panopto.com/Panopto/Pages/Viewer.aspx?id={video_id}"
            html_content += f'<h1><a href="{video_link}">{formatted_name} <span style="color: #1155cc;">(link)</span></a></h1>\n'
        else:
            html_content += f"<h1>{formatted_name}</h1>\n"
        
        # Extract all topics from all batches and sort chronologically
        all_topics = []
        
        for i, (batch, summary) in enumerate(zip(batches, batch_summaries), 1):
            # Extract topics from the summary with their timestamps
            topics = self._extract_topics_from_summary(summary, video_id, transcript_data)
            
            # # Update timestamps to better match content if transcript data available
            if transcript_data:
                topics = self._update_topic_timestamps(topics, transcript_data)
            
            # Add batch info to each topic
            for topic in topics:
                topic['batch_index'] = i
                topic['batch'] = batch
            
            all_topics.extend(topics)
        
        # Sort all topics by timestamp_seconds
        all_topics.sort(key=lambda x: x['timestamp_seconds'] if x['timestamp_seconds'] is not None else float('inf'))
        
        # Generate HTML content with ordered list
        html_content += "<ol>\n"
        
        for topic_info in all_topics:
            topic = topic_info['topic']
            speaker = topic_info['speaker']
            content = topic_info['content']
            
            # Format the list item with exact original structure
            html_content += '<li><h3 class="topic-heading">'
            
            if topic_info['video_link'] and topic_info['timestamp_seconds'] is not None:
                # Verify timestamp format
                seconds = topic_info['timestamp_seconds']
                corrected_timestamp = self._seconds_to_time_str(seconds)
                
                html_content += f'<a href="{topic_info["video_link"]}" class="topic-link">'
                html_content += f'{topic} - {speaker} <span style="color: #1155cc;">({corrected_timestamp})</span></a>'
            else:
                html_content += f'{topic} - {speaker} ({topic_info["timestamp"]})'
            
            html_content += '</h3>\n'
            html_content += f'<div class="topic-content">{content}</div></li>\n'
        
        html_content += "</ol>\n</body>\n</html>"
        
        return html_content
    
    def _generate_meeting_markdown(self, batches: List[List[Dict]], batch_summaries: List[str], 
                              video_id: Optional[str], transcript_data: List[Dict], 
                              metadata: Dict = None) -> str:
        """Generate markdown for meeting summaries - matching original format."""
        md_lines = []
        
        # Get meeting name from metadata
        meeting_name = metadata.get('meeting_name') if metadata else None
        if not meeting_name and hasattr(self, 'metadata'):
            meeting_name = self.metadata.get('meeting_name')
        if meeting_name:
            formatted_name = self._format_meeting_name(meeting_name)
        else:
            formatted_name = 'Meeting Summary'
        
        # Add title
        if video_id:
            video_link = f"https://mit.hosted.panopto.com/Panopto/Pages/Viewer.aspx?id={video_id}"
            md_lines.append(f"# [{formatted_name}]({video_link})\n")
        else:
            md_lines.append(f"# {formatted_name}\n")
        
        # Extract all topics from all batches
        all_topics = []
        
        for i, (batch, summary) in enumerate(zip(batches, batch_summaries), 1):
            # Extract topics from the summary
            topics = self._extract_topics_from_summary(summary, video_id, transcript_data)
            
            # Update timestamps if transcript data available
            if transcript_data:
                topics = self._update_topic_timestamps(topics, transcript_data)
            
            all_topics.extend(topics)
        
        # Sort all topics by timestamp
        all_topics.sort(key=lambda x: x['timestamp_seconds'] if x['timestamp_seconds'] is not None else float('inf'))
        
        # Process each topic
        for topic_info in all_topics:
            topic = topic_info['topic']
            speaker = topic_info['speaker']
            content = topic_info['content']
            
            # Verify and correct timestamp
            if topic_info['timestamp_seconds'] is not None:
                corrected_timestamp = self._seconds_to_time_str(topic_info['timestamp_seconds'])
            else:
                corrected_timestamp = topic_info['timestamp']
            
            # Format the topic header
            if topic_info['video_link']:
                md_lines.append(f"**{topic} - {speaker}** [({corrected_timestamp})]({topic_info['video_link']})")
            else:
                md_lines.append(f"**{topic} - {speaker}** ({corrected_timestamp})")
            
            # Add the content
            md_lines.append(f"{content}\n")
        
        return '\n'.join(md_lines)
    

    def _create_timestamp_reference(self, batch_entries: List[Dict]) -> str:
        """
        Create a simple timestamp reference for GPT to use.
        
        This is the original approach - just list all timestamps chronologically
        for each speaker with a text preview.
        
        Args:
            batch_entries: List of transcript entries with speaker, timestamp, and text
            
        Returns:
            Formatted string with timestamps organized by speaker
        """
        # Create a mapping of speaker names to ALL their timestamps for this batch
        speaker_timestamps = {}
        for entry in batch_entries:
            speaker = entry['name']
            if speaker not in speaker_timestamps:
                speaker_timestamps[speaker] = []
            
            # Add this timestamp to the list for this speaker
            speaker_timestamps[speaker].append({
                'seconds': entry['seconds'],
                'time_str': entry['time_str'],
                'text': entry['text'][:100]  # Include a snippet of text for context
            })
        
        # Prepare the timestamp reference for the model
        timestamp_reference = "SPEAKER TIMESTAMPS (DO NOT MODIFY THESE):\n"
        
        # Sort speakers alphabetically for consistency
        for speaker in sorted(speaker_timestamps.keys()):
            timestamps = speaker_timestamps[speaker]
            
            # Sort timestamps chronologically
            sorted_timestamps = sorted(timestamps, key=lambda x: x['seconds'])
            
            # Include all timestamps for the speaker with context snippets
            timestamp_reference += f"\n{speaker}:\n"
            for i, ts in enumerate(sorted_timestamps, 1):
                timestamp_reference += f"  {i}. {ts['time_str']} - '{ts['text']}...'\n"
        
        return timestamp_reference

    def _update_topic_timestamps(self, topics: List[Dict], transcript_data: List[Dict]) -> List[Dict]:
        """Update topic timestamps to better match actual content."""
        return topics

    def _find_best_timestamp_match(self, topic_content: str, speaker: str, transcript_data: List[Dict]) -> Optional[Dict]:
        """Find the best timestamp match for a topic in the transcript."""
        # Filter entries by speaker
        speaker_entries = [entry for entry in transcript_data if entry['name'] == speaker]
        
        if not speaker_entries:
            return None
        
        # If NLTK is available, use advanced matching
        if NLTK_AVAILABLE:
            try:
                from nltk.tokenize import word_tokenize
                from nltk.corpus import stopwords
                
                # Preprocess topic content
                stop_words = set(stopwords.words('english'))
                topic_words = [w.lower() for w in word_tokenize(topic_content) if w.isalnum() and w.lower() not in stop_words]
                
                if not topic_words:
                    return speaker_entries[0]
                
                # Calculate similarity scores
                best_match = None
                highest_score = -1
                
                for entry in speaker_entries:
                    entry_words = [w.lower() for w in word_tokenize(entry['text']) if w.isalnum() and w.lower() not in stop_words]
                    
                    if not entry_words:
                        continue
                    
                    # Calculate word overlap
                    overlap = len(set(topic_words) & set(entry_words))
                    
                    # Calculate similarity using TF-IDF if available
                    try:
                        vectorizer = TfidfVectorizer()
                        tfidf_matrix = vectorizer.fit_transform([topic_content, entry['text']])
                        similarity = cosine_similarity(tfidf_matrix[0:1], tfidf_matrix[1:2])[0][0]
                    except:
                        similarity = overlap / max(len(topic_words), 1)
                    
                    # Combined score
                    score = (0.7 * similarity) + (0.3 * (overlap / max(len(topic_words), 1)))
                    
                    if score > highest_score:
                        highest_score = score
                        best_match = entry
                
                # Return best match if good enough, otherwise first entry
                return best_match if best_match and highest_score > 0.1 else speaker_entries[0]
                
            except Exception as e:
                print(f"      Warning: Advanced matching failed: {e}")
        
        # Basic fallback: return first entry for speaker
        return speaker_entries[0]
    def _extract_topics_from_summary(self, summary: str, video_id: Optional[str], 
                                     transcript_data: List[Dict]) -> List[Dict]:
        """Extract topics from a batch summary."""
        # Pattern to match: **Topic - Speaker** (H:MM:SS):
        pattern = r'\*\*(.+?)\s+-\s+(.+?)\*\*\s*\((\d+:\d{2}:\d{2})\)\s*:'
        
        topics = []
        matches = list(re.finditer(pattern, summary))
        
        for idx, match in enumerate(matches):
            topic = match.group(1).strip()
            speaker = match.group(2).strip()
            timestamp = match.group(3)
            
            # Convert timestamp to seconds
            timestamp_seconds = self._time_to_seconds(timestamp)
            
            # Create video link if video_id provided
            video_link = None
            if video_id:
                video_link = f'https://mit.hosted.panopto.com/Panopto/Pages/Viewer.aspx?id={video_id}&start={timestamp_seconds}'
            
            # Extract content (between this match and the next)
            start_pos = match.end()
            end_pos = matches[idx + 1].start() if idx + 1 < len(matches) else len(summary)
            content = summary[start_pos:end_pos].strip()
            
            topics.append({
                'topic': topic,
                'speaker': speaker,
                'timestamp': timestamp,
                'timestamp_seconds': timestamp_seconds,
                'video_link': video_link,
                'content': content
            })
        
        return topics
    
    def _generate_meeting_summaries_fallback(self) -> Tuple[str, str]:
        """Generate fallback meeting summaries when AI is not available."""
        html = """<!DOCTYPE html>
<html>
<head>
<title>Meeting Summaries</title>
<style>
body { font-family: Arial, sans-serif; margin: 20px; }
</style>
</head>
<body>
<h1>Meeting Summaries</h1>
<p>AI summaries are not available. Please configure OpenAI API key in .env file.</p>
</body>
</html>"""
        
        md = "# Meeting Summaries\n\nAI summaries are not available. Please configure OpenAI API key."
        
        return html, md
    
    def _convert_bold_formatting(self, html_files: Dict[str, str]) -> None:
        """Convert markdown-style bold formatting to HTML bold tags."""
        pattern = r'\*\*([^*]+?)\*\*'
        
        for file_type, file_path in html_files.items():
            if file_path.endswith('.html'):
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                    
                    # Replace **text** with <b>text</b> in content sections
                    updated_content = re.sub(pattern, r'<b>\1</b>', content)
                    
                    with open(file_path, 'w', encoding='utf-8') as f:
                        f.write(updated_content)
                    
                    print(f"    Converted bold formatting in: {file_path}")
                    
                except Exception as e:
                    print(f"    Warning: Failed to convert bold formatting in {file_path}: {e}")
            
            elif file_path.endswith('.md'):
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                    
                    # For markdown files, convert to <b> tags but preserve headers
                    lines = content.split('\n')
                    for i in range(len(lines)):
                        # Skip lines that are headers or already have special formatting
                        if not lines[i].startswith('#') and not lines[i].startswith('**') or ' - ' not in lines[i][:40]:
                            lines[i] = re.sub(pattern, r'<b>\1</b>', lines[i])
                    
                    updated_content = '\n'.join(lines)
                    
                    with open(file_path, 'w', encoding='utf-8') as f:
                        f.write(updated_content)
                    
                    print(f"    Converted bold formatting in: {file_path}")
                    
                except Exception as e:
                    print(f"    Warning: Failed to convert bold formatting in {file_path}: {e}")
    
    def _format_meeting_name(self, raw_name: str) -> str:
        """Format meeting name for display in HTML/Markdown."""
        import re
        
        # Replace underscores with spaces
        formatted = raw_name.replace('_', ' ')
        
        # Fix timestamp formatting (e.g., "4.00pm" -> "4:00pm")
        formatted = re.sub(r'(?<=\d)\.(\d{2})(am|pm)', r':\1\2', formatted)
        
        return formatted